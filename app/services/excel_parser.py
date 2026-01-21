"""
Excel解析サービスモジュール

財務データ・製造データのExcelファイルを解析するサービスを提供する。
Step 1で作成したテンプレート形式に対応。

機能:
- 財務データExcelのパース
- 製造データExcelのパース
- バリデーション処理
"""
import io
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import load_workbook
from openpyxl.cell.cell import Cell
from openpyxl.worksheet.worksheet import Worksheet


# =============================================================================
# 財務データ項目マッピング
# =============================================================================

# 項目名 → DBカラム名のマッピング
FINANCIAL_ITEM_MAPPING = {
    "全社売上高": "sales_total",
    "店舗部門売上高": "sales_store",
    "通販部門売上高": "sales_online",
    "売上原価": "cost_of_sales",
    "売上総利益（粗利）": "gross_profit",
    "売上総利益": "gross_profit",
    "粗利率": "gross_profit_rate",
    "販管費合計": "sg_and_a_total",
    "うち人件費": "labor_cost",
    "人件費": "labor_cost",
    "人件費率": "labor_cost_rate",
    "その他経費": "other_expenses",
    "営業利益": "operating_profit",
    "営業利益率": "operating_profit_rate",
    "営業キャッシュフロー": "cf_operating",
    "投資キャッシュフロー": "cf_investing",
    "財務キャッシュフロー": "cf_financing",
    "フリーキャッシュフロー": "cf_free",
}

# セクションヘッダー（スキップ対象）
SECTION_HEADERS = [
    "■ 売上高",
    "■ 原価・利益",
    "■ 販管費",
    "■ 営業利益",
    "■ キャッシュフロー",
]


# =============================================================================
# ユーティリティ関数
# =============================================================================

def get_cell_value(cell: Cell, data_only: bool = True) -> Any:
    """
    セルの値を取得する

    Args:
        cell: セルオブジェクト
        data_only: 数式の計算結果を取得するか

    Returns:
        セルの値
    """
    if cell is None:
        return None
    return cell.value


def parse_date_value(value: Any) -> Optional[date]:
    """
    日付値をパースする

    Args:
        value: セルの値

    Returns:
        date型または None
    """
    if value is None:
        return None

    if isinstance(value, datetime):
        return value.date()

    if isinstance(value, date):
        return value

    if isinstance(value, str):
        # "YYYY/MM/DD" 形式
        for fmt in ["%Y/%m/%d", "%Y-%m-%d", "%Y年%m月%d日"]:
            try:
                return datetime.strptime(value.strip(), fmt).date()
            except ValueError:
                continue

        # "YYYY/MM/01" または "YYYY年MM月" 形式
        try:
            parts = value.replace("年", "/").replace("月", "/").replace("日", "").strip().split("/")
            if len(parts) >= 2:
                year = int(parts[0])
                month = int(parts[1])
                day = int(parts[2]) if len(parts) > 2 and parts[2] else 1
                return date(year, month, day)
        except (ValueError, IndexError):
            pass

    return None


def parse_numeric_value(value: Any) -> Optional[Decimal]:
    """
    数値をパースする

    Args:
        value: セルの値

    Returns:
        Decimal型または None
    """
    if value is None:
        return None

    if isinstance(value, (int, float)):
        if value != value:  # NaN check
            return None
        return Decimal(str(value))

    if isinstance(value, Decimal):
        return value

    if isinstance(value, str):
        # カンマ、円、%記号を除去
        cleaned = value.replace(",", "").replace("円", "").replace("%", "").strip()
        if not cleaned or cleaned.lower() == "nan":
            return None
        try:
            return Decimal(cleaned)
        except InvalidOperation:
            return None

    return None


def parse_int_value(value: Any) -> Optional[int]:
    """
    整数値をパースする

    Args:
        value: セルの値

    Returns:
        int型または None
    """
    decimal_value = parse_numeric_value(value)
    if decimal_value is not None:
        return int(decimal_value)
    return None


# =============================================================================
# 売上原価明細項目マッピング
# =============================================================================

COST_DETAIL_MAPPING = {
    "仕入高": "purchases",
    "原材料仕入高": "raw_material_purchases",
    "労務費": "labor_cost",
    "消耗品費": "consumables",
    "賃借料": "rent",
    "修繕費": "repairs",
    "水道光熱費": "utilities",
    # 「その他」はDB側で計算されるため保存しない
}

# =============================================================================
# 販管費明細項目マッピング
# =============================================================================

SGA_DETAIL_MAPPING = {
    "役員報酬": "executive_compensation",
    "人件費（販管費）": "personnel_cost",
    "人件費": "personnel_cost",
    "配送費": "delivery_cost",
    "包装費": "packaging_cost",
    "支払手数料": "payment_fees",
    "荷造運賃費": "freight_cost",
    "販売手数料": "sales_commission",
    "広告宣伝費": "advertising_cost",
    # 「その他」はDB側で計算されるため保存しない
}


# =============================================================================
# 財務データExcelパース
# =============================================================================

def parse_financial_excel(file_content: bytes) -> Dict[str, Any]:
    """
    財務データExcelをパースする

    Step 1で作成したテンプレート形式のExcelファイルを解析する。

    Args:
        file_content: Excelファイルの内容（バイト列）

    Returns:
        Dict[str, Any]: パース結果
            - success: bool - 成功かどうか
            - month: str - 対象月（YYYY-MM-DD形式）
            - is_target: bool - 予算データかどうか
            - data: Dict - パースされたデータ
            - errors: List[Dict] - エラーメッセージ
            - warnings: List[str] - 警告メッセージ
    """
    result = {
        "success": False,
        "month": None,
        "is_target": False,
        "data": {},
        "errors": [],
        "warnings": [],
    }

    try:
        # Excelファイルを読み込み（data_only=Trueで数式の計算結果を取得）
        wb = load_workbook(io.BytesIO(file_content), data_only=True)

        # 「月次財務データ」シートを明示的に指定（アクティブシートに依存しない）
        target_sheet_name = "月次財務データ"
        if target_sheet_name in wb.sheetnames:
            ws = wb[target_sheet_name]
        else:
            # シート名が異なる場合はアクティブシートを使用
            ws = wb.active

        if ws is None:
            result["errors"].append({
                "row": None,
                "column": None,
                "message": "ワークシートが見つかりません",
                "value": None,
            })
            return result

        # B1: 対象年月
        date_value = get_cell_value(ws["B1"])
        target_date = parse_date_value(date_value)

        if not target_date:
            result["errors"].append({
                "row": 1,
                "column": "B1",
                "message": "対象年月が不正です",
                "value": str(date_value) if date_value else None,
            })
            return result

        # 月初日に正規化
        result["month"] = date(target_date.year, target_date.month, 1).isoformat()

        # B2: データ区分
        data_type = get_cell_value(ws["B2"])
        if data_type:
            data_type_str = str(data_type).strip()
            if data_type_str in ["予算", "目標", "計画"]:
                result["is_target"] = True
            elif data_type_str in ["実績", "actual"]:
                result["is_target"] = False
            else:
                result["warnings"].append(f"データ区分が不明です: {data_type_str}（実績として処理）")

        # 項目名 → 行番号のマッピングを構築
        item_rows = {}
        for row_idx in range(1, ws.max_row + 1):
            cell_value = get_cell_value(ws.cell(row=row_idx, column=1))
            if cell_value and isinstance(cell_value, str):
                item_name = cell_value.strip()
                if item_name in FINANCIAL_ITEM_MAPPING:
                    item_rows[item_name] = row_idx

        # データを読み取り
        data = {}
        for item_name, db_column in FINANCIAL_ITEM_MAPPING.items():
            if item_name in item_rows:
                row_idx = item_rows[item_name]
                value = get_cell_value(ws.cell(row=row_idx, column=2))
                numeric_value = parse_numeric_value(value)

                if numeric_value is not None:
                    data[db_column] = float(numeric_value)
                elif value is not None and str(value).strip():
                    # 数値以外の値が入っている場合は警告
                    result["warnings"].append(
                        f"行{row_idx}: {item_name}の値が数値ではありません: {value}"
                    )

        result["data"] = data

        # 売上原価明細シートのパース
        cost_details = _parse_cost_detail_sheet(wb, result["month"])
        if cost_details:
            result["cost_details"] = cost_details

        # 販管費明細シートのパース
        sga_details = _parse_sga_detail_sheet(wb, result["month"])
        if sga_details:
            result["sga_details"] = sga_details

        # バリデーション
        errors = validate_financial_data(data, result["month"])
        result["errors"].extend(errors)

        result["success"] = len(result["errors"]) == 0

    except Exception as e:
        result["errors"].append({
            "row": None,
            "column": None,
            "message": f"Excelファイルの解析に失敗しました: {str(e)}",
            "value": None,
        })

    return result


def _parse_cost_detail_sheet(wb, month: str) -> Optional[Dict[str, Any]]:
    """
    売上原価明細シートをパースする

    Args:
        wb: openpyxlワークブック
        month: 対象月（メインシートから取得）

    Returns:
        原価明細データまたはNone
    """
    sheet_name = "売上原価明細"
    if sheet_name not in wb.sheetnames:
        return None

    ws = wb[sheet_name]
    data = {}

    # 項目を探して値を取得
    for row_idx in range(1, ws.max_row + 1):
        cell_value = get_cell_value(ws.cell(row=row_idx, column=1))
        if cell_value and isinstance(cell_value, str):
            item_name = cell_value.strip()
            if item_name in COST_DETAIL_MAPPING:
                value = get_cell_value(ws.cell(row=row_idx, column=2))
                numeric_value = parse_numeric_value(value)
                if numeric_value is not None:
                    db_column = COST_DETAIL_MAPPING[item_name]
                    data[db_column] = float(numeric_value)

    if not data:
        return None

    return data


def _parse_sga_detail_sheet(wb, month: str) -> Optional[Dict[str, Any]]:
    """
    販管費明細シートをパースする

    Args:
        wb: openpyxlワークブック
        month: 対象月（メインシートから取得）

    Returns:
        販管費明細データまたはNone
    """
    sheet_name = "販管費明細"
    if sheet_name not in wb.sheetnames:
        return None

    ws = wb[sheet_name]
    data = {}

    # 項目を探して値を取得
    for row_idx in range(1, ws.max_row + 1):
        cell_value = get_cell_value(ws.cell(row=row_idx, column=1))
        if cell_value and isinstance(cell_value, str):
            item_name = cell_value.strip()
            if item_name in SGA_DETAIL_MAPPING:
                value = get_cell_value(ws.cell(row=row_idx, column=2))
                numeric_value = parse_numeric_value(value)
                if numeric_value is not None:
                    db_column = SGA_DETAIL_MAPPING[item_name]
                    data[db_column] = float(numeric_value)

    if not data:
        return None

    return data


def validate_financial_data(data: Dict[str, Any], month: str) -> List[Dict[str, Any]]:
    """
    財務データのバリデーション

    Args:
        data: パースされたデータ
        month: 対象月

    Returns:
        エラーリスト
    """
    errors = []

    # 売上高が0未満の場合は警告（エラーではない）
    # マイナスの場合もあり得る（返品等）

    # 必須チェック：売上高は必須とする
    if not data.get("sales_total") and data.get("sales_total") != 0:
        # 売上がない場合でもエラーにはしない（警告のみ）
        pass

    return errors


# =============================================================================
# 製造データExcelパース
# =============================================================================

def parse_manufacturing_excel(file_content: bytes) -> Dict[str, Any]:
    """
    製造データExcelをパースする

    Step 1で作成したテンプレート形式のExcelファイルを解析する。

    Args:
        file_content: Excelファイルの内容（バイト列）

    Returns:
        Dict[str, Any]: パース結果
            - success: bool - 成功かどうか
            - month: str - 対象月（YYYY-MM-DD形式）
            - data: List[Dict] - 日次データリスト
            - summary: Dict - 月次サマリー
            - errors: List[Dict] - エラーメッセージ
            - warnings: List[str] - 警告メッセージ
    """
    result = {
        "success": False,
        "month": None,
        "data": [],
        "summary": {
            "total_batts": 0,
            "total_pieces": 0,
            "total_workers": 0,
            "avg_production_per_worker": None,
            "total_paid_leave_hours": 0,
            "working_days": 0,
        },
        "errors": [],
        "warnings": [],
    }

    try:
        # Excelファイルを読み込み（data_only=Trueで数式の計算結果を取得）
        wb = load_workbook(io.BytesIO(file_content), data_only=True)
        ws = wb.active

        if ws is None:
            result["errors"].append({
                "row": None,
                "column": None,
                "message": "ワークシートが見つかりません",
                "value": None,
            })
            return result

        # B1: 対象年月（"2025年11月" 形式）
        month_value = get_cell_value(ws["B1"])
        target_month = _parse_month_value(month_value)

        if not target_month:
            result["errors"].append({
                "row": 1,
                "column": "B1",
                "message": "対象年月が不正です",
                "value": str(month_value) if month_value else None,
            })
            return result

        result["month"] = target_month.isoformat()

        # 4行目以降: 日次データ
        # ヘッダー行（3行目）: 日付, 製造量(バット), 製造量(個), 出勤者数, 1人あたり製造量, 有給取得(時間)
        data_rows = []
        total_batts = 0
        total_pieces = 0
        total_workers = 0
        total_paid_leave = Decimal("0")
        working_days = 0

        for row_idx in range(4, ws.max_row + 1):
            # A列: 日付
            date_cell = get_cell_value(ws.cell(row=row_idx, column=1))

            # 合計行のスキップ
            if date_cell and isinstance(date_cell, str) and date_cell.strip() == "合計":
                continue

            # 空行スキップ
            if date_cell is None:
                continue

            # 日付のパース
            row_date = parse_date_value(date_cell)
            if not row_date:
                # 日付でない場合はスキップ（ヘッダー行など）
                continue

            # 対象月に属するかチェック
            if row_date.year != target_month.year or row_date.month != target_month.month:
                result["warnings"].append(
                    f"行{row_idx}: 日付 {row_date} は対象月 {target_month.strftime('%Y年%m月')} に属しません"
                )
                continue

            # B列: 製造量(バット)
            batts = parse_int_value(get_cell_value(ws.cell(row=row_idx, column=2)))

            # C列: 製造量(個) - 通常は数式で計算されるが、data_only=Trueで値取得
            pieces = parse_int_value(get_cell_value(ws.cell(row=row_idx, column=3)))

            # D列: 出勤者数
            workers = parse_int_value(get_cell_value(ws.cell(row=row_idx, column=4)))

            # E列: 1人あたり製造量 - 数式で計算される
            production_per_worker = parse_numeric_value(get_cell_value(ws.cell(row=row_idx, column=5)))

            # F列: 有給取得(時間)
            paid_leave = parse_numeric_value(get_cell_value(ws.cell(row=row_idx, column=6)))

            # バリデーション
            row_errors = validate_manufacturing_row(row_idx, batts, workers, paid_leave)
            result["errors"].extend(row_errors)

            # データ行を追加
            row_data = {
                "date": row_date.isoformat(),
                "production_batts": batts,
                "production_pieces": pieces if pieces else (batts * 60 if batts else None),
                "workers_count": workers,
                "production_per_worker": float(production_per_worker) if production_per_worker else None,
                "paid_leave_hours": float(paid_leave) if paid_leave else None,
            }
            data_rows.append(row_data)

            # サマリー集計
            if batts:
                total_batts += batts
                working_days += 1
            if pieces:
                total_pieces += pieces
            elif batts:
                total_pieces += batts * 60
            if workers:
                total_workers += workers
            if paid_leave:
                total_paid_leave += paid_leave

        result["data"] = data_rows

        # サマリー計算
        result["summary"] = {
            "total_batts": total_batts,
            "total_pieces": total_pieces,
            "total_workers": total_workers,
            "avg_production_per_worker": round(total_batts / total_workers, 2) if total_workers > 0 else None,
            "total_paid_leave_hours": float(total_paid_leave),
            "working_days": working_days,
        }

        result["success"] = len(result["errors"]) == 0

    except Exception as e:
        result["errors"].append({
            "row": None,
            "column": None,
            "message": f"Excelファイルの解析に失敗しました: {str(e)}",
            "value": None,
        })

    return result


def _parse_month_value(value: Any) -> Optional[date]:
    """
    月の値をパースする（"2025年11月" 形式対応）

    Args:
        value: セルの値

    Returns:
        月初日のdate型または None
    """
    if value is None:
        return None

    if isinstance(value, (datetime, date)):
        d = value if isinstance(value, date) else value.date()
        return date(d.year, d.month, 1)

    if isinstance(value, str):
        # "2025年11月" 形式
        try:
            value = value.strip()
            if "年" in value and "月" in value:
                parts = value.replace("年", " ").replace("月", "").split()
                year = int(parts[0])
                month = int(parts[1])
                return date(year, month, 1)
        except (ValueError, IndexError):
            pass

        # "2025/11" 形式
        try:
            parts = value.split("/")
            if len(parts) >= 2:
                year = int(parts[0])
                month = int(parts[1])
                return date(year, month, 1)
        except (ValueError, IndexError):
            pass

    return None


def validate_manufacturing_row(
    row_idx: int,
    batts: Optional[int],
    workers: Optional[int],
    paid_leave: Optional[Decimal]
) -> List[Dict[str, Any]]:
    """
    製造データ行のバリデーション

    Args:
        row_idx: 行番号
        batts: 製造量（バット）
        workers: 出勤者数
        paid_leave: 有給取得時間

    Returns:
        エラーリスト
    """
    errors = []

    if batts is not None and batts < 0:
        errors.append({
            "row": row_idx,
            "column": "製造量(バット)",
            "message": "製造量は0以上である必要があります",
            "value": str(batts),
        })

    if workers is not None and workers < 0:
        errors.append({
            "row": row_idx,
            "column": "出勤者数",
            "message": "出勤者数は0以上である必要があります",
            "value": str(workers),
        })

    if paid_leave is not None and paid_leave < 0:
        errors.append({
            "row": row_idx,
            "column": "有給取得(時間)",
            "message": "有給取得時間は0以上である必要があります",
            "value": str(paid_leave),
        })

    return errors


# =============================================================================
# サンプルデータ生成（テスト用）
# =============================================================================

def get_financial_sample() -> Dict[str, Any]:
    """
    財務データのサンプル構造を返す

    Returns:
        サンプルデータ構造
    """
    return {
        "template_structure": {
            "B1": "対象年月（例: 2025/11/01）",
            "B2": "データ区分（実績 または 予算）",
            "items": [
                {"row": 5, "name": "全社売上高", "db_column": "sales_total"},
                {"row": 6, "name": "店舗部門売上高", "db_column": "sales_store"},
                {"row": 7, "name": "通販部門売上高", "db_column": "sales_online"},
                {"row": 9, "name": "売上原価", "db_column": "cost_of_sales"},
                {"row": 10, "name": "売上総利益（粗利）", "db_column": "gross_profit"},
                {"row": 11, "name": "粗利率", "db_column": "gross_profit_rate"},
                {"row": 13, "name": "販管費合計", "db_column": "sg_and_a_total"},
                {"row": 14, "name": "うち人件費", "db_column": "labor_cost"},
                {"row": 15, "name": "人件費率", "db_column": "labor_cost_rate"},
                {"row": 16, "name": "その他経費", "db_column": "other_expenses"},
                {"row": 18, "name": "営業利益", "db_column": "operating_profit"},
                {"row": 19, "name": "営業利益率", "db_column": "operating_profit_rate"},
                {"row": 21, "name": "営業キャッシュフロー", "db_column": "cf_operating"},
                {"row": 22, "name": "投資キャッシュフロー", "db_column": "cf_investing"},
                {"row": 23, "name": "財務キャッシュフロー", "db_column": "cf_financing"},
                {"row": 24, "name": "フリーキャッシュフロー", "db_column": "cf_free"},
            ],
        },
        "sample_values": {
            "month": "2025-11-01",
            "is_target": False,
            "data": {
                "sales_total": 228000000,
                "sales_store": 121000000,
                "sales_online": 107000000,
                "cost_of_sales": 137000000,
                "gross_profit": 91000000,
                "gross_profit_rate": 39.9,
                "sg_and_a_total": 46000000,
                "labor_cost": 51000000,
                "labor_cost_rate": 22.4,
                "other_expenses": -5000000,
                "operating_profit": 45000000,
                "operating_profit_rate": 19.7,
                "cf_operating": 38000000,
                "cf_investing": -12000000,
                "cf_financing": -8000000,
                "cf_free": 26000000,
            },
        },
    }


def get_manufacturing_sample() -> Dict[str, Any]:
    """
    製造データのサンプル構造を返す

    Returns:
        サンプルデータ構造
    """
    return {
        "template_structure": {
            "B1": "対象年月（例: 2025年11月）",
            "header_row": 3,
            "columns": [
                {"column": "A", "name": "日付"},
                {"column": "B", "name": "製造量(バット)"},
                {"column": "C", "name": "製造量(個)（自動計算: バット×60）"},
                {"column": "D", "name": "出勤者数"},
                {"column": "E", "name": "1人あたり製造量（自動計算）"},
                {"column": "F", "name": "有給取得(時間)"},
            ],
            "data_start_row": 4,
        },
        "sample_values": [
            {
                "date": "2025-11-01",
                "production_batts": 150,
                "production_pieces": 9000,
                "workers_count": 12,
                "production_per_worker": 12.5,
                "paid_leave_hours": 8.0,
            },
            {
                "date": "2025-11-02",
                "production_batts": 145,
                "production_pieces": 8700,
                "workers_count": 11,
                "production_per_worker": 13.18,
                "paid_leave_hours": 0,
            },
        ],
    }

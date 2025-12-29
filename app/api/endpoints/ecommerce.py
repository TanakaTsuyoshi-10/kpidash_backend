"""
通販分析APIエンドポイントモジュール

通販チャネル別・商品別・顧客別実績およびHPアクセス数の
取得・登録APIを提供する。
"""
import csv
import io
import zipfile
from datetime import date
from typing import Optional

from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File, status
from fastapi.responses import StreamingResponse
from supabase import Client

from app.api.deps import get_current_user, get_supabase_admin
from app.schemas.kpi import User
from app.schemas.ecommerce import (
    ChannelSummaryResponse,
    ProductSummaryResponse,
    CustomerSummaryResponse,
    WebsiteStatsResponse,
    TrendResponse,
    EcommerceUploadResponse,
    EcommerceBulkUploadResponse,
)
from app.services.ecommerce_service import (
    get_channel_summary,
    get_product_summary,
    get_customer_summary,
    get_website_stats,
    get_ecommerce_trend,
    import_channel_data,
    import_product_data,
    import_customer_data,
    import_website_data,
)
from app.services.metrics import get_fiscal_year


router = APIRouter(prefix="/ecommerce", tags=["ecommerce"])


# =============================================================================
# テンプレート定義
# =============================================================================

# 商品リスト（https://www.gyo-za.co.jp/products/list より）
# 個数バリエーションを含む
PRODUCT_LIST = [
    # 生餃子（個数バリエーションあり）
    "ぎょうざ 50個入",
    "ぎょうざ 40個入",
    "しょうが入りぎょうざ 50個入",
    "しょうが入りぎょうざ 40個入",
    # たれ・スープ類
    "焼きぎょうざのたれ",
    "日向夏のたれ",
    "水ぎょうざのスープ",
    "ぎょうざ鍋すうぷ",
    "ぎょうざ鍋の素 塩味",
    "ぎょうざのたれ 柚子こしょう",
    "ぎょうざのたれ みそ",
]

TEMPLATES = {
    "channel": {
        "filename": "ecommerce_channel_template.csv",
        "description": "チャネル別実績テンプレート",
        "headers": ["チャネル", "売上高", "購入者数"],
        "sample_data": [
            ["EC", "", ""],
            ["電話", "", ""],
            ["FAX", "", ""],
            ["店舗受付", "", ""],
        ],
    },
    "product": {
        "filename": "ecommerce_product_template.csv",
        "description": "商品別実績テンプレート",
        "headers": ["商品名", "商品カテゴリ", "売上高", "販売数量"],
        "sample_data": [
            ["ぎょうざ（冷凍）", "冷凍食品", "", ""],
            ["しょうが入ぎょうざ（冷凍）", "冷凍食品", "", ""],
        ],
    },
    "customer": {
        "filename": "ecommerce_customer_template.csv",
        "description": "顧客別実績テンプレート",
        "headers": ["新規顧客数", "リピーター数"],
        "sample_data": [["", ""]],
    },
    "website": {
        "filename": "ecommerce_website_template.csv",
        "description": "HPアクセス数テンプレート",
        "headers": ["ページビュー数", "ユニークビジター数", "セッション数"],
        "sample_data": [["", "", ""]],
    },
}


# =============================================================================
# テンプレートダウンロードエンドポイント
# =============================================================================

@router.get(
    "/template/{data_type}",
    summary="テンプレートダウンロード",
    description="通販データ登録用のCSVテンプレートをダウンロードする。",
)
async def download_template(
    data_type: str,
    current_user: User = Depends(get_current_user),
):
    """
    テンプレートファイルをダウンロード

    Args:
        data_type: データタイプ（channel, product, customer, website, all）

    Returns:
        CSVファイルまたはZIPファイル
    """
    if data_type == "all":
        # 全テンプレートをZIPで返す
        zip_buffer = io.BytesIO()
        with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zf:
            for key, template in TEMPLATES.items():
                csv_content = create_csv_content(template)
                zf.writestr(template["filename"], csv_content)

        zip_buffer.seek(0)
        return StreamingResponse(
            zip_buffer,
            media_type="application/zip",
            headers={
                "Content-Disposition": "attachment; filename=ecommerce_templates.zip"
            },
        )

    if data_type not in TEMPLATES:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"不正なdata_type: {data_type}。有効な値: channel, product, customer, website, all"
        )

    template = TEMPLATES[data_type]
    csv_content = create_csv_content(template)

    return StreamingResponse(
        io.StringIO(csv_content),
        media_type="text/csv; charset=utf-8",
        headers={
            "Content-Disposition": f"attachment; filename={template['filename']}"
        },
    )


def create_csv_content(template: dict) -> str:
    """テンプレートCSVコンテンツを生成"""
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(template["headers"])
    for row in template["sample_data"]:
        writer.writerow(row)
    return output.getvalue()


# =============================================================================
# Excelテンプレートダウンロードエンドポイント
# =============================================================================

@router.get(
    "/template-excel",
    summary="Excelテンプレートダウンロード",
    description="通販データ登録用のExcelテンプレートをダウンロードする。1ファイルに全データを含む。",
)
async def download_excel_template(
    current_user: User = Depends(get_current_user),
):
    """
    Excelテンプレートファイルをダウンロード

    1つのExcelファイルに以下のシートを含む:
    - 基本情報（対象月入力欄）
    - チャネル別実績
    - 商品別実績
    - 顧客別実績
    - HPアクセス数

    Returns:
        Excelファイル
    """
    import openpyxl
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()

    # スタイル定義
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font_white = Font(bold=True, size=11, color="FFFFFF")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    title_font = Font(bold=True, size=14)
    section_font = Font(bold=True, size=12, color="2F5496")

    # シート1: 入力シート（すべてのデータを1シートに）
    ws = wb.active
    ws.title = "通販実績データ"

    current_row = 1

    # タイトル
    ws.cell(row=current_row, column=1, value="通販部門 月次実績データ入力シート")
    ws.cell(row=current_row, column=1).font = title_font
    current_row += 2

    # 対象月入力欄
    ws.cell(row=current_row, column=1, value="対象月")
    ws.cell(row=current_row, column=1).font = header_font
    ws.cell(row=current_row, column=2, value="")  # 入力欄
    ws.cell(row=current_row, column=2).border = thin_border
    ws.cell(row=current_row, column=3, value="← YYYY-MM-DD形式で入力（例: 2025-11-01）")
    ws.cell(row=current_row, column=3).font = Font(italic=True, color="808080")
    current_row += 3

    # セクション1: チャネル別実績
    ws.cell(row=current_row, column=1, value="■ チャネル別実績")
    ws.cell(row=current_row, column=1).font = section_font
    current_row += 1

    channel_headers = ["チャネル", "売上高", "購入者数"]
    for col, header in enumerate(channel_headers, 1):
        cell = ws.cell(row=current_row, column=col, value=header)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')
    current_row += 1

    channels = ["EC", "電話", "FAX", "店舗受付"]
    for channel in channels:
        ws.cell(row=current_row, column=1, value=channel).border = thin_border
        ws.cell(row=current_row, column=2, value="").border = thin_border
        ws.cell(row=current_row, column=3, value="").border = thin_border
        current_row += 1

    current_row += 2

    # セクション2: 商品別実績
    ws.cell(row=current_row, column=1, value="■ 商品別実績")
    ws.cell(row=current_row, column=1).font = section_font
    current_row += 1

    product_headers = ["商品名", "売上高", "販売数量"]
    for col, header in enumerate(product_headers, 1):
        cell = ws.cell(row=current_row, column=col, value=header)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')
    current_row += 1

    for product in PRODUCT_LIST:
        ws.cell(row=current_row, column=1, value=product).border = thin_border
        ws.cell(row=current_row, column=2, value="").border = thin_border
        ws.cell(row=current_row, column=3, value="").border = thin_border
        current_row += 1

    current_row += 2

    # セクション3: 顧客別実績
    ws.cell(row=current_row, column=1, value="■ 顧客別実績")
    ws.cell(row=current_row, column=1).font = section_font
    current_row += 1

    customer_headers = ["新規顧客数", "リピーター数"]
    for col, header in enumerate(customer_headers, 1):
        cell = ws.cell(row=current_row, column=col, value=header)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')
    current_row += 1

    ws.cell(row=current_row, column=1, value="").border = thin_border
    ws.cell(row=current_row, column=2, value="").border = thin_border
    current_row += 3

    # セクション4: HPアクセス数
    ws.cell(row=current_row, column=1, value="■ HPアクセス数")
    ws.cell(row=current_row, column=1).font = section_font
    current_row += 1

    website_headers = ["ページビュー数", "ユニークビジター数", "セッション数"]
    for col, header in enumerate(website_headers, 1):
        cell = ws.cell(row=current_row, column=col, value=header)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')
    current_row += 1

    ws.cell(row=current_row, column=1, value="").border = thin_border
    ws.cell(row=current_row, column=2, value="").border = thin_border
    ws.cell(row=current_row, column=3, value="").border = thin_border

    # 列幅調整
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 40

    # Excelファイルをバイトストリームに保存
    excel_buffer = io.BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)

    return StreamingResponse(
        excel_buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": "attachment; filename=ecommerce_template.xlsx"
        },
    )


# =============================================================================
# データアップロードエンドポイント
# =============================================================================

@router.post(
    "/upload",
    response_model=EcommerceUploadResponse,
    summary="通販データアップロード",
    description="CSV/Excelファイルから通販データを登録する。",
)
async def upload_ecommerce_data(
    file: UploadFile = File(..., description="CSV/Excelファイル"),
    data_type: str = Query(..., description="データタイプ: channel, product, customer, website"),
    month: date = Query(..., description="対象月（YYYY-MM-DD形式）"),
    current_user: User = Depends(get_current_user),
    supabase: Client = Depends(get_supabase_admin),
) -> EcommerceUploadResponse:
    """
    通販データをアップロード

    Args:
        file: アップロードファイル
        data_type: データタイプ
        month: 対象月

    Returns:
        EcommerceUploadResponse: アップロード結果
    """
    if data_type not in TEMPLATES:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"不正なdata_type: {data_type}"
        )

    # ファイル形式チェック
    filename = file.filename.lower()
    if not (filename.endswith(".csv") or filename.endswith(".xlsx") or filename.endswith(".xls")):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="CSV形式またはExcel形式のファイルをアップロードしてください"
        )

    try:
        content = await file.read()

        # CSVの場合
        if filename.endswith(".csv"):
            # エンコーディング検出
            try:
                text = content.decode("utf-8-sig")
            except UnicodeDecodeError:
                text = content.decode("cp932")

            reader = csv.DictReader(io.StringIO(text))
            records = list(reader)

        # Excelの場合
        else:
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(content))
            ws = wb.active
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail="ファイルにデータがありません"
                )
            headers = [str(h) if h else "" for h in rows[0]]
            records = []
            for row in rows[1:]:
                record = {}
                for i, val in enumerate(row):
                    if i < len(headers):
                        record[headers[i]] = val
                records.append(record)

        if not records:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="ファイルにデータがありません"
            )

        # データタイプに応じてインポート
        if data_type == "channel":
            result = await import_channel_data(supabase, month, records)
        elif data_type == "product":
            result = await import_product_data(supabase, month, records)
        elif data_type == "customer":
            result = await import_customer_data(supabase, month, records)
        elif data_type == "website":
            result = await import_website_data(supabase, month, records)
        else:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"不正なdata_type: {data_type}"
            )

        return EcommerceUploadResponse(
            success=True,
            message="データのアップロードが完了しました",
            data_type=data_type,
            month=month.isoformat(),
            records_processed=len(records),
            records_created=result.get("created", 0),
            records_updated=result.get("updated", 0),
        )

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
            detail=f"ファイルの処理中にエラーが発生しました: {str(e)}"
        )


# =============================================================================
# Excel一括アップロードエンドポイント
# =============================================================================

@router.post(
    "/upload-excel",
    response_model=EcommerceBulkUploadResponse,
    summary="Excel一括アップロード",
    description="""
    Excelテンプレートから通販データを一括登録する。

    テンプレートの構造:
    - 行3・列B: 対象月（YYYY-MM-DD形式）
    - 行8-11: チャネル別データ（EC, 電話, FAX, 店舗受付）
    - 行16-26: 商品別データ
    - 行31: 顧客別データ
    - 行36: HPアクセスデータ
    """,
)
async def upload_excel_bulk(
    file: UploadFile = File(..., description="Excelファイル"),
    current_user: User = Depends(get_current_user),
    supabase: Client = Depends(get_supabase_admin),
) -> EcommerceBulkUploadResponse:
    """
    Excelテンプレートから一括アップロード

    Args:
        file: Excelテンプレートファイル

    Returns:
        EcommerceBulkUploadResponse: アップロード結果
    """
    import openpyxl
    from datetime import datetime as dt

    filename = file.filename.lower()
    if not (filename.endswith(".xlsx") or filename.endswith(".xls")):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Excel形式のファイル（.xlsx, .xls）をアップロードしてください"
        )

    try:
        content = await file.read()
        # data_only=True で数式の計算結果を読み込む
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
        ws = wb.active

        # 対象月を取得（行3、列B）
        month_value = ws.cell(row=3, column=2).value
        if not month_value:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="対象月が入力されていません（行3・列B）"
            )

        # 日付パース
        if isinstance(month_value, str):
            try:
                target_month = dt.strptime(month_value.strip(), "%Y-%m-%d").date()
            except ValueError:
                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail="対象月の形式が不正です。YYYY-MM-DD形式で入力してください"
                )
        elif hasattr(month_value, 'date'):
            target_month = month_value.date() if hasattr(month_value, 'date') else month_value
        elif isinstance(month_value, date):
            target_month = month_value
        else:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="対象月の形式が不正です"
            )

        # 月の1日に正規化
        target_month = date(target_month.year, target_month.month, 1)

        channel_count = 0
        product_count = 0
        customer_count = 0
        website_count = 0

        def parse_numeric(value):
            """数値をパースする。数式文字列や無効な値はNoneを返す"""
            if value is None:
                return None
            if isinstance(value, (int, float)):
                return value
            if isinstance(value, str):
                # 数式文字列（=で始まる）はスキップ
                if value.startswith('=') or value.startswith('+'):
                    return None
                # 数値文字列をパース
                try:
                    return float(value.replace(',', ''))
                except ValueError:
                    return None
            return None

        # チャネル別データ取得（行8-11）
        channel_records = []
        for row_num in range(8, 12):
            channel_name = ws.cell(row=row_num, column=1).value
            sales = parse_numeric(ws.cell(row=row_num, column=2).value)
            buyers = parse_numeric(ws.cell(row=row_num, column=3).value)

            if channel_name and (sales is not None or buyers is not None):
                channel_records.append({
                    "チャネル": str(channel_name),
                    "売上高": sales,
                    "購入者数": int(buyers) if buyers is not None else None,
                })

        if channel_records:
            result = await import_channel_data(supabase, target_month, channel_records)
            channel_count = result.get("created", 0) + result.get("updated", 0)

        # 商品別データ取得（行16-26）
        product_records = []
        for row_num in range(16, 27):
            product_name = ws.cell(row=row_num, column=1).value
            sales = parse_numeric(ws.cell(row=row_num, column=2).value)
            quantity = parse_numeric(ws.cell(row=row_num, column=3).value)

            if product_name and (sales is not None or quantity is not None):
                product_records.append({
                    "商品名": str(product_name),
                    "商品カテゴリ": None,
                    "売上高": sales,
                    "販売数量": int(quantity) if quantity is not None else None,
                })

        if product_records:
            result = await import_product_data(supabase, target_month, product_records)
            product_count = result.get("created", 0) + result.get("updated", 0)

        # 顧客別データ取得（行31）
        new_customers = parse_numeric(ws.cell(row=31, column=1).value)
        repeat_customers = parse_numeric(ws.cell(row=31, column=2).value)

        if new_customers is not None or repeat_customers is not None:
            customer_records = [{
                "新規顧客数": int(new_customers) if new_customers is not None else None,
                "リピーター数": int(repeat_customers) if repeat_customers is not None else None,
            }]
            result = await import_customer_data(supabase, target_month, customer_records)
            customer_count = result.get("created", 0) + result.get("updated", 0)

        # HPアクセスデータ取得（行36）
        page_views = parse_numeric(ws.cell(row=36, column=1).value)
        unique_visitors = parse_numeric(ws.cell(row=36, column=2).value)
        sessions = parse_numeric(ws.cell(row=36, column=3).value)

        if page_views is not None or unique_visitors is not None or sessions is not None:
            website_records = [{
                "ページビュー数": int(page_views) if page_views is not None else None,
                "ユニークビジター数": int(unique_visitors) if unique_visitors is not None else None,
                "セッション数": int(sessions) if sessions is not None else None,
            }]
            result = await import_website_data(supabase, target_month, website_records)
            website_count = result.get("created", 0) + result.get("updated", 0)

        total_count = channel_count + product_count + customer_count + website_count
        if total_count == 0:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="アップロードするデータがありません。テンプレートにデータを入力してください"
            )

        return EcommerceBulkUploadResponse(
            success=True,
            message="データの一括アップロードが完了しました",
            month=target_month.isoformat(),
            channel_records=channel_count,
            product_records=product_count,
            customer_records=customer_count,
            website_records=website_count,
        )

    except HTTPException:
        raise
    except Exception as e:
        import traceback
        print(f"Excel upload error: {str(e)}")
        print(traceback.format_exc())
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
            detail=f"ファイルの処理中にエラーが発生しました: {str(e)}"
        )


# =============================================================================
# チャネル別実績エンドポイント
# =============================================================================

@router.get(
    "/channel-summary",
    response_model=ChannelSummaryResponse,
    summary="チャネル別実績取得",
    description="""
    チャネル別（EC、電話、FAX、店舗受付）の売上実績を取得する。

    - 各チャネルの売上高・購入者数・客単価
    - 前年比・前々年比（累計モード時）
    - 合計データ

    ## パラメータ
    - month: 対象月（YYYY-MM-DD形式）
    - period_type: 期間タイプ（monthly: 単月, cumulative: 9月〜対象月の累計）
    """,
)
async def channel_summary(
    month: date = Query(..., description="対象月（YYYY-MM-DD形式）"),
    period_type: str = Query("monthly", description="期間タイプ（monthly/cumulative）"),
    current_user: User = Depends(get_current_user),
    supabase: Client = Depends(get_supabase_admin),
) -> ChannelSummaryResponse:
    """チャネル別実績を取得"""
    if period_type not in ["monthly", "cumulative"]:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="period_typeは'monthly'または'cumulative'を指定してください"
        )

    try:
        result = await get_channel_summary(supabase, month, period_type)
        return ChannelSummaryResponse(**result)
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"チャネル別実績の取得に失敗しました: {str(e)}"
        )


# =============================================================================
# 商品別実績エンドポイント
# =============================================================================

@router.get(
    "/product-summary",
    response_model=ProductSummaryResponse,
    summary="商品別実績取得",
    description="""
    商品別の売上実績を取得する。

    - 商品名・カテゴリ・売上高・販売数量
    - 前年比・前々年比
    - 売上高順でソート

    ## パラメータ
    - month: 対象月（YYYY-MM-DD形式）
    - period_type: 期間タイプ（monthly/cumulative）
    - limit: 取得件数（デフォルト: 20）
    """,
)
async def product_summary(
    month: date = Query(..., description="対象月（YYYY-MM-DD形式）"),
    period_type: str = Query("monthly", description="期間タイプ（monthly/cumulative）"),
    limit: int = Query(20, ge=1, le=100, description="取得件数"),
    current_user: User = Depends(get_current_user),
    supabase: Client = Depends(get_supabase_admin),
) -> ProductSummaryResponse:
    """商品別実績を取得"""
    if period_type not in ["monthly", "cumulative"]:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="period_typeは'monthly'または'cumulative'を指定してください"
        )

    try:
        result = await get_product_summary(supabase, month, period_type, limit)
        return ProductSummaryResponse(**result)
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"商品別実績の取得に失敗しました: {str(e)}"
        )


# =============================================================================
# 顧客別実績エンドポイント
# =============================================================================

@router.get(
    "/customer-summary",
    response_model=CustomerSummaryResponse,
    summary="顧客別実績取得",
    description="""
    顧客別の統計データを取得する。

    - 新規顧客数・リピーター数・合計顧客数
    - リピート率
    - 前年比・前々年比

    ## パラメータ
    - month: 対象月（YYYY-MM-DD形式）
    - period_type: 期間タイプ（monthly/cumulative）
    """,
)
async def customer_summary(
    month: date = Query(..., description="対象月（YYYY-MM-DD形式）"),
    period_type: str = Query("monthly", description="期間タイプ（monthly/cumulative）"),
    current_user: User = Depends(get_current_user),
    supabase: Client = Depends(get_supabase_admin),
) -> CustomerSummaryResponse:
    """顧客別実績を取得"""
    if period_type not in ["monthly", "cumulative"]:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="period_typeは'monthly'または'cumulative'を指定してください"
        )

    try:
        result = await get_customer_summary(supabase, month, period_type)
        return CustomerSummaryResponse(**result)
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"顧客別実績の取得に失敗しました: {str(e)}"
        )


# =============================================================================
# HPアクセス数エンドポイント
# =============================================================================

@router.get(
    "/website-stats",
    response_model=WebsiteStatsResponse,
    summary="HPアクセス数取得",
    description="""
    HPアクセス統計データを取得する。

    - ページビュー数・ユニークビジター数・セッション数
    - 前年比・前々年比

    ## パラメータ
    - month: 対象月（YYYY-MM-DD形式）
    - period_type: 期間タイプ（monthly/cumulative）
    """,
)
async def website_stats(
    month: date = Query(..., description="対象月（YYYY-MM-DD形式）"),
    period_type: str = Query("monthly", description="期間タイプ（monthly/cumulative）"),
    current_user: User = Depends(get_current_user),
    supabase: Client = Depends(get_supabase_admin),
) -> WebsiteStatsResponse:
    """HPアクセス数を取得"""
    if period_type not in ["monthly", "cumulative"]:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="period_typeは'monthly'または'cumulative'を指定してください"
        )

    try:
        result = await get_website_stats(supabase, month, period_type)
        return WebsiteStatsResponse(**result)
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"HPアクセス数の取得に失敗しました: {str(e)}"
        )


# =============================================================================
# 推移データエンドポイント（グラフ用）
# =============================================================================

@router.get(
    "/trend",
    response_model=TrendResponse,
    summary="推移データ取得",
    description="""
    グラフ表示用の月次推移データを取得する。

    ## metric（指標タイプ）
    - channel_sales: チャネル別売上推移
    - product_sales: 商品別売上推移（上位10商品）
    - customers: 顧客数推移
    - website: HPアクセス数推移

    ## 会計年度
    - 9月〜翌8月を1年度とする
    - fiscal_year省略時は現在の会計年度
    """,
)
async def trend_data(
    metric: str = Query(..., description="指標タイプ: channel_sales, product_sales, customers, website"),
    fiscal_year: Optional[int] = Query(None, description="会計年度（省略時は現在）"),
    current_user: User = Depends(get_current_user),
    supabase: Client = Depends(get_supabase_admin),
) -> TrendResponse:
    """推移データを取得"""
    valid_metrics = ["channel_sales", "product_sales", "customers", "website"]
    if metric not in valid_metrics:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"不正なmetric: {metric}。有効な値: {', '.join(valid_metrics)}"
        )

    try:
        result = await get_ecommerce_trend(supabase, metric, fiscal_year)
        return TrendResponse(**result)
    except ValueError as e:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=str(e)
        )
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"推移データの取得に失敗しました: {str(e)}"
        )

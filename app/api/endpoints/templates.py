"""
テンプレートダウンロードAPIエンドポイントモジュール

財務データ・製造データ入力用のExcelテンプレートをダウンロードするAPIを提供する。
"""
import calendar
import io
from datetime import date, datetime
from pathlib import Path
from typing import Optional, List, Dict, Any

from fastapi import APIRouter, Depends, HTTPException, Query, status
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from supabase import Client

from app.api.deps import get_supabase_admin


router = APIRouter(prefix="/templates", tags=["templates"])

# スタイル定義
HEADER_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
HEADER_FONT = Font(bold=True)
CENTER_ALIGNMENT = Alignment(horizontal="center", vertical="center")
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def generate_financial_template(year: int, month: int) -> io.BytesIO:
    """
    財務データ入力用Excelテンプレートを生成する（詳細版）

    Args:
        year: 対象年
        month: 対象月

    Returns:
        Excelファイルのバイトストリーム
    """
    wb = Workbook()

    # ========== シート1: 月次財務データ（基本） ==========
    ws = wb.active
    ws.title = "月次財務データ"

    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 15
    ws.column_dimensions["C"].width = 8

    ws["A1"] = "対象年月"
    ws["B1"] = f"{year}/{month:02d}/01"
    ws["A2"] = "データ区分"
    ws["B2"] = "実績"

    items = [
        ("■ 売上高", None, None, True),
        ("全社売上高", None, "円", False),
        ("店舗部門売上高", None, "円", False),
        ("通販部門売上高", None, "円", False),
        ("■ 原価・利益", None, None, True),
        ("売上原価", None, "円", False),
        ("売上総利益（粗利）", "=B6-B10", "円", False),
        ("粗利率", "=IF(B6>0,B11/B6*100,0)", "%", False),
        ("■ 販管費", None, None, True),
        ("販管費合計", None, "円", False),
        ("■ 営業利益", None, None, True),
        ("営業利益", "=B11-B14", "円", False),
        ("営業利益率", "=IF(B6>0,B16/B6*100,0)", "%", False),
        ("■ キャッシュフロー", None, None, True),
        ("営業キャッシュフロー", None, "円", False),
        ("投資キャッシュフロー", None, "円", False),
        ("財務キャッシュフロー", None, "円", False),
        ("フリーキャッシュフロー", "=B19+B20", "円", False),
    ]

    row = 4
    for item_name, formula, unit, is_section in items:
        ws[f"A{row}"] = item_name
        if formula:
            ws[f"B{row}"] = formula
        if unit:
            ws[f"C{row}"] = unit
        if is_section:
            ws[f"A{row}"].font = HEADER_FONT
            ws[f"A{row}"].fill = HEADER_FILL
            ws[f"B{row}"].fill = HEADER_FILL
            ws[f"C{row}"].fill = HEADER_FILL
        row += 1

    # ========== シート2: 売上原価明細 ==========
    ws2 = wb.create_sheet("売上原価明細")
    ws2.column_dimensions["A"].width = 20
    ws2.column_dimensions["B"].width = 15
    ws2.column_dimensions["C"].width = 8

    ws2["A1"] = "対象年月"
    ws2["B1"] = f"{year}/{month:02d}/01"

    cost_items = [
        ("■ 売上原価内訳", None, None, True),
        ("仕入高", None, "円", False),
        ("原材料仕入高", None, "円", False),
        ("労務費", None, "円", False),
        ("消耗品費", None, "円", False),
        ("賃借料", None, "円", False),
        ("修繕費", None, "円", False),
        ("水道光熱費", None, "円", False),
        ("その他", "=月次財務データ!B10-SUM(B5:B11)", "円", False),
        ("売上原価合計（参照）", "=月次財務データ!B10", "円", True),
    ]

    row = 3
    for item_name, formula, unit, is_section in cost_items:
        ws2[f"A{row}"] = item_name
        if formula:
            ws2[f"B{row}"] = formula
        if unit:
            ws2[f"C{row}"] = unit
        if is_section:
            ws2[f"A{row}"].font = HEADER_FONT
            ws2[f"A{row}"].fill = HEADER_FILL
            ws2[f"B{row}"].fill = HEADER_FILL
            ws2[f"C{row}"].fill = HEADER_FILL
        row += 1

    # ========== シート3: 販管費明細 ==========
    ws3 = wb.create_sheet("販管費明細")
    ws3.column_dimensions["A"].width = 20
    ws3.column_dimensions["B"].width = 15
    ws3.column_dimensions["C"].width = 8

    ws3["A1"] = "対象年月"
    ws3["B1"] = f"{year}/{month:02d}/01"

    sga_items = [
        ("■ 販管費内訳", None, None, True),
        ("役員報酬", None, "円", False),
        ("人件費（販管費）", None, "円", False),
        ("配送費", None, "円", False),
        ("包装費", None, "円", False),
        ("支払手数料", None, "円", False),
        ("荷造運賃費", None, "円", False),
        ("販売手数料", None, "円", False),
        ("広告宣伝費", None, "円", False),
        ("その他", "=月次財務データ!B14-SUM(B5:B12)", "円", False),
        ("販管費合計（参照）", "=月次財務データ!B14", "円", True),
    ]

    row = 3
    for item_name, formula, unit, is_section in sga_items:
        ws3[f"A{row}"] = item_name
        if formula:
            ws3[f"B{row}"] = formula
        if unit:
            ws3[f"C{row}"] = unit
        if is_section:
            ws3[f"A{row}"].font = HEADER_FONT
            ws3[f"A{row}"].fill = HEADER_FILL
            ws3[f"B{row}"].fill = HEADER_FILL
            ws3[f"C{row}"].fill = HEADER_FILL
        row += 1

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def generate_store_pl_template(
    year: int,
    month: int,
    stores: List[Dict[str, Any]]
) -> io.BytesIO:
    """
    店舗別収支入力用Excelテンプレートを生成する

    Args:
        year: 対象年
        month: 対象月
        stores: 店舗一覧（id, code, nameを含む辞書のリスト）

    Returns:
        Excelファイルのバイトストリーム
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "店舗別収支"

    # 列幅設定
    ws.column_dimensions["A"].width = 20  # 店舗名用に幅を広げる
    ws.column_dimensions["B"].width = 15
    ws.column_dimensions["C"].width = 15
    ws.column_dimensions["D"].width = 15
    ws.column_dimensions["E"].width = 15
    ws.column_dimensions["F"].width = 15
    ws.column_dimensions["G"].width = 15
    ws.column_dimensions["H"].width = 15
    ws.column_dimensions["I"].width = 15
    ws.column_dimensions["J"].width = 15

    # ヘッダー情報
    ws["A1"] = "対象年月"
    ws["B1"] = f"{year}/{month:02d}/01"

    # テーブルヘッダー
    headers = [
        "店舗名",
        "売上高",
        "売上原価",
        "売上総利益",
        "販管費合計",
        "営業利益",
        "人件費",
        "地代家賃",
        "賃借料",
        "水道光熱費",
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER_ALIGNMENT
        cell.border = THIN_BORDER

    # 店舗データ行を動的に生成
    data_start_row = 4
    for i, store in enumerate(stores):
        row = data_start_row + i
        # 店舗名を表示
        ws.cell(row=row, column=1, value=store.get("name", "")).border = THIN_BORDER
        for col in range(2, 11):
            ws.cell(row=row, column=col).border = THIN_BORDER

        # 売上総利益 = 売上高 - 売上原価
        ws.cell(row=row, column=4, value=f"=B{row}-C{row}").border = THIN_BORDER
        # 営業利益 = 売上総利益 - 販管費
        ws.cell(row=row, column=6, value=f"=D{row}-E{row}").border = THIN_BORDER

    # 合計行
    store_count = len(stores)
    sum_row = data_start_row + store_count
    last_data_row = sum_row - 1

    ws.cell(row=sum_row, column=1, value="合計").font = HEADER_FONT
    ws.cell(row=sum_row, column=1).fill = HEADER_FILL
    ws.cell(row=sum_row, column=1).border = THIN_BORDER

    for col in range(2, 11):
        col_letter = get_column_letter(col)
        if col in [4, 6]:  # 計算式列はスキップ
            continue
        cell = ws.cell(
            row=sum_row,
            column=col,
            value=f"=SUM({col_letter}{data_start_row}:{col_letter}{last_data_row})"
        )
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER

    # 売上総利益合計
    ws.cell(row=sum_row, column=4, value=f"=B{sum_row}-C{sum_row}").font = HEADER_FONT
    ws.cell(row=sum_row, column=4).fill = HEADER_FILL
    ws.cell(row=sum_row, column=4).border = THIN_BORDER

    # 営業利益合計
    ws.cell(row=sum_row, column=6, value=f"=D{sum_row}-E{sum_row}").font = HEADER_FONT
    ws.cell(row=sum_row, column=6).fill = HEADER_FILL
    ws.cell(row=sum_row, column=6).border = THIN_BORDER

    # 説明シートを追加
    ws2 = wb.create_sheet("入力説明")
    ws2.column_dimensions["A"].width = 15
    ws2.column_dimensions["B"].width = 50

    ws2["A1"] = "項目名"
    ws2["B1"] = "説明"
    ws2["A1"].font = HEADER_FONT
    ws2["B1"].font = HEADER_FONT

    explanations = [
        ("店舗名", "店舗名は自動で設定されています（編集しないでください）"),
        ("売上高", "店舗の月間売上高"),
        ("売上原価", "店舗の月間売上原価"),
        ("売上総利益", "自動計算（売上高 - 売上原価）"),
        ("販管費合計", "店舗の月間販管費合計"),
        ("営業利益", "自動計算（売上総利益 - 販管費）"),
        ("人件費", "店舗の人件費"),
        ("地代家賃", "店舗の地代家賃"),
        ("賃借料", "店舗のリース・賃借料"),
        ("水道光熱費", "店舗の水道光熱費"),
        ("その他", "販管費合計との差額として自動計算"),
    ]

    for i, (item, desc) in enumerate(explanations, 2):
        ws2[f"A{i}"] = item
        ws2[f"B{i}"] = desc

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def generate_manufacturing_template(year: int, month: int) -> io.BytesIO:
    """
    製造データ入力用Excelテンプレートを生成する

    Args:
        year: 対象年
        month: 対象月

    Returns:
        Excelファイルのバイトストリーム
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "日次製造データ"

    # 列幅設定
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 15
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 18
    ws.column_dimensions["F"].width = 15

    # ヘッダー情報
    ws["A1"] = "対象年月"
    ws["B1"] = f"{year}年{month:02d}月"

    # テーブルヘッダー
    headers = [
        "日付",
        "製造量(バット)",
        "製造量(個)",
        "出勤者数",
        "1人あたり製造量",
        "有給取得(時間)",
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER_ALIGNMENT
        cell.border = THIN_BORDER

    # 月の日数を取得
    _, days_in_month = calendar.monthrange(year, month)

    # データ行を生成
    data_start_row = 4
    for day in range(1, days_in_month + 1):
        row = data_start_row + day - 1
        day_date = date(year, month, day)

        ws.cell(row=row, column=1, value=day_date).border = THIN_BORDER
        ws.cell(row=row, column=2).border = THIN_BORDER
        ws.cell(row=row, column=3, value=f"=B{row}*60").border = THIN_BORDER
        ws.cell(row=row, column=4).border = THIN_BORDER
        ws.cell(row=row, column=5, value=f"=IF(D{row}>0,B{row}/D{row},0)").border = THIN_BORDER
        ws.cell(row=row, column=6).border = THIN_BORDER

    # 合計行
    sum_row = data_start_row + days_in_month
    ws.cell(row=sum_row, column=1, value="合計").font = HEADER_FONT
    ws.cell(row=sum_row, column=1).fill = HEADER_FILL
    ws.cell(row=sum_row, column=1).border = THIN_BORDER

    for col in range(2, 7):
        col_letter = get_column_letter(col)
        if col == 5:
            formula = f"=IF(D{sum_row}>0,B{sum_row}/D{sum_row},0)"
        else:
            formula = f"=SUM({col_letter}{data_start_row}:{col_letter}{sum_row-1})"

        cell = ws.cell(row=sum_row, column=col, value=formula)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER

    # バイトストリームに保存
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


# =============================================================================
# エンドポイント
# =============================================================================

@router.get(
    "/financial",
    summary="財務データテンプレートダウンロード",
    description="""
    財務データ入力用のExcelテンプレートをダウンロードする。

    ## パラメータ
    - year: 対象年（省略時は現在の年）
    - month: 対象月（省略時は現在の月）

    ## 出力
    - Excel形式（.xlsx）
    - 売上高、原価・利益、販管費、営業利益、キャッシュフロー等の入力項目
    - 自動計算式付き（粗利率、営業利益率など）
    """,
)
async def download_financial_template(
    year: Optional[int] = Query(None, description="対象年"),
    month: Optional[int] = Query(None, ge=1, le=12, description="対象月"),
):
    """
    財務データテンプレートをダウンロード

    Args:
        year: 対象年
        month: 対象月

    Returns:
        Excelファイル
    """
    now = datetime.now()
    year = year or now.year
    month = month or now.month

    try:
        output = generate_financial_template(year, month)
        filename = f"financial_template_{year}{month:02d}.xlsx"

        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f'attachment; filename="{filename}"'
            },
        )
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"テンプレート生成に失敗しました: {str(e)}",
        )


@router.get(
    "/manufacturing",
    summary="製造データテンプレートダウンロード",
    description="""
    製造データ入力用のExcelテンプレートをダウンロードする。

    ## パラメータ
    - year: 対象年（省略時は現在の年）
    - month: 対象月（省略時は現在の月）

    ## 出力
    - Excel形式（.xlsx）
    - 日次の製造量、出勤者数、有給取得時間の入力項目
    - 自動計算式付き（製造量(個)、1人あたり製造量など）
    """,
)
async def download_manufacturing_template(
    year: Optional[int] = Query(None, description="対象年"),
    month: Optional[int] = Query(None, ge=1, le=12, description="対象月"),
):
    """
    製造データテンプレートをダウンロード

    Args:
        year: 対象年
        month: 対象月

    Returns:
        Excelファイル
    """
    now = datetime.now()
    year = year or now.year
    month = month or now.month

    try:
        output = generate_manufacturing_template(year, month)
        filename = f"manufacturing_template_{year}{month:02d}.xlsx"

        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f'attachment; filename="{filename}"'
            },
        )
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"テンプレート生成に失敗しました: {str(e)}",
        )


@router.get(
    "/store-pl",
    summary="店舗別収支テンプレートダウンロード",
    description="""
    店舗別収支入力用のExcelテンプレートをダウンロードする。

    ## パラメータ
    - year: 対象年（省略時は現在の年）
    - month: 対象月（省略時は現在の月）
    - department_slug: 部門スラッグ（デフォルト: store）

    ## 出力
    - Excel形式（.xlsx）
    - DBに登録されている全店舗分の行を動的に生成
    - 店舗別の売上高、売上原価、販管費、営業利益の入力項目
    - 販管費明細（人件費、地代家賃、賃借料、水道光熱費）
    - 自動計算式付き（売上総利益、営業利益）
    """,
)
async def download_store_pl_template(
    year: Optional[int] = Query(None, description="対象年"),
    month: Optional[int] = Query(None, ge=1, le=12, description="対象月"),
    department_slug: str = Query("store", description="部門スラッグ"),
    supabase: Client = Depends(get_supabase_admin),
):
    """
    店舗別収支テンプレートをダウンロード

    Args:
        year: 対象年
        month: 対象月
        department_slug: 部門スラッグ
        supabase: Supabaseクライアント

    Returns:
        Excelファイル
    """
    now = datetime.now()
    year = year or now.year
    month = month or now.month

    try:
        # 部門IDを取得
        dept_response = supabase.table("departments").select(
            "id"
        ).eq("slug", department_slug).execute()

        if not dept_response.data:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail=f"部門が見つかりません: {department_slug}",
            )

        department_id = dept_response.data[0]["id"]

        # 店舗一覧を取得
        stores_response = supabase.table("segments").select(
            "id, code, name"
        ).eq("department_id", department_id).order("code").execute()

        stores = stores_response.data or []

        if not stores:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="店舗が登録されていません",
            )

        output = generate_store_pl_template(year, month, stores)
        filename = f"store_pl_template_{year}{month:02d}.xlsx"

        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f'attachment; filename="{filename}"'
            },
        )
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"テンプレート生成に失敗しました: {str(e)}",
        )

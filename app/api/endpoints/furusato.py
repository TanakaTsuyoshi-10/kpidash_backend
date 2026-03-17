"""
ふるさと納税分析APIエンドポイントモジュール

ふるさと納税の販売実績・リピート情報・返品苦情・口コミの
取得・Excelアップロード機能を提供する。
"""
import io
from datetime import date

from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File, status
from supabase import Client

from app.api.deps import get_current_user, get_supabase_admin
from app.schemas.kpi import User
from app.schemas.furusato import FurusatoSummaryResponse, FurusatoUploadResponse
from app.services.furusato_service import get_furusato_summary, import_furusato_data


router = APIRouter(prefix="/furusato", tags=["furusato"])


# =============================================================================
# サマリー取得エンドポイント
# =============================================================================

@router.get(
    "/summary",
    response_model=FurusatoSummaryResponse,
    summary="ふるさと納税サマリー取得",
    description="""
    ふるさと納税の各種データを取得する。

    - 販売実績（在庫数・注文数・売上高・エリア別注文数）
    - リピート情報（新規注文者・EC購入経験者・複数回購入者）
    - 返品・苦情情報（再送数・苦情数）
    - 口コミ情報（ポジティブ・ネガティブ）

    ## パラメータ
    - month: 対象月（YYYY-MM-DD形式）
    - period_type: 期間タイプ（monthly: 単月, cumulative: 9月〜対象月の累計）
    """,
)
async def furusato_summary(
    month: date = Query(..., description="対象月（YYYY-MM-DD形式）"),
    period_type: str = Query("monthly", description="期間タイプ（monthly/cumulative）"),
    current_user: User = Depends(get_current_user),
    supabase: Client = Depends(get_supabase_admin),
) -> FurusatoSummaryResponse:
    """ふるさと納税サマリーを取得"""
    if period_type not in ["monthly", "cumulative"]:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="period_typeは'monthly'または'cumulative'を指定してください"
        )

    try:
        result = await get_furusato_summary(supabase, month, period_type)
        return FurusatoSummaryResponse(**result)
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"ふるさと納税サマリーの取得に失敗しました: {str(e)}"
        )


# =============================================================================
# Excelアップロードエンドポイント
# =============================================================================

@router.post(
    "/upload-excel",
    response_model=FurusatoUploadResponse,
    summary="ふるさと納税Excelアップロード",
    description="""
    ふるさと納税管理表（.xlsm/.xlsx）の「集計表」シートからデータを取り込む。

    - Row1/Col4のExcel日付値から対象月を判定
    - 各行の合計列（Col J = index 10）を読み取り
    - 販売実績・リピート情報・返品苦情・口コミデータを保存
    """,
)
async def upload_furusato_excel(
    file: UploadFile = File(..., description="ふるさと納税管理表Excelファイル"),
    current_user: User = Depends(get_current_user),
    supabase: Client = Depends(get_supabase_admin),
) -> FurusatoUploadResponse:
    """ふるさと納税Excelをアップロード"""
    import openpyxl
    from datetime import datetime as dt

    filename = file.filename.lower()
    if not (filename.endswith(".xlsx") or filename.endswith(".xls") or filename.endswith(".xlsm")):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Excel形式のファイル（.xlsx, .xls, .xlsm）をアップロードしてください"
        )

    try:
        content = await file.read()
        # data_only=True で数式の計算結果を読み込む
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)

        # 「集計表」シートを探す
        sheet_name = None
        for name in wb.sheetnames:
            if "集計表" in name:
                sheet_name = name
                break

        if not sheet_name:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="「集計表」シートが見つかりません"
            )

        ws = wb[sheet_name]

        # 対象月を取得（Row1, Col4 = D1）
        month_value = ws.cell(row=1, column=4).value
        if month_value is None or month_value == "":
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"対象月が入力されていません（Row1・Col D）。シート名: {sheet_name}"
            )

        # 日付パース
        target_month = None
        if isinstance(month_value, str):
            # 文字列の場合、複数フォーマットを試行
            for fmt in ["%Y-%m-%d", "%Y/%m/%d", "%Y年%m月", "%Y年%m月%d日"]:
                try:
                    target_month = dt.strptime(month_value.strip(), fmt).date()
                    break
                except ValueError:
                    continue
            if target_month is None:
                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail=f"対象月の形式が不正です（文字列: '{month_value}'）"
                )
        elif isinstance(month_value, (int, float)):
            # Excelシリアル日付（例: 46054 = 2026-02-01）
            from openpyxl.utils.datetime import from_excel
            try:
                target_month = from_excel(month_value).date()
            except Exception:
                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail=f"対象月の形式が不正です（数値: {month_value}）"
                )
        elif hasattr(month_value, 'date'):
            target_month = month_value.date()
        elif isinstance(month_value, date):
            target_month = month_value
        else:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"対象月の形式が不正です（型: {type(month_value).__name__}, 値: {month_value}）"
            )

        # 月の1日に正規化
        target_month = date(target_month.year, target_month.month, 1)

        def parse_numeric(value):
            """数値をパースする"""
            if value is None:
                return None
            if isinstance(value, (int, float)):
                return value
            if isinstance(value, str):
                if value.startswith('=') or value.startswith('+'):
                    return None
                try:
                    return float(value.replace(',', ''))
                except ValueError:
                    return None
            return None

        def get_val(row, col=10):
            """指定行・列の値を取得（デフォルトはCol J = 10）"""
            return parse_numeric(ws.cell(row=row, column=col).value)

        def get_weekly(row):
            """第1〜5週の値を取得（Col E=5, F=6, G=7, H=8, I=9）"""
            return [parse_numeric(ws.cell(row=row, column=c).value) for c in range(5, 10)]

        def get_comment(row, col=12):
            """指定行・列のコメントを取得（Col L = 12）"""
            val = ws.cell(row=row, column=col).value
            return str(val) if val else None

        # データ取り込み
        data = {
            # 販売実績
            "inventory": int(get_val(4)) if get_val(4) is not None else None,
            "orders": int(get_val(5)) if get_val(5) is not None else None,
            "sales": get_val(8),
            "unit_price": parse_numeric(ws.cell(row=9, column=5).value),  # 単価はCol E
            "orders_kyushu": int(get_val(11)) if get_val(11) is not None else None,
            "orders_chugoku_shikoku": int(get_val(12)) if get_val(12) is not None else None,
            "orders_kansai": int(get_val(13)) if get_val(13) is not None else None,
            "orders_kanto": int(get_val(14)) if get_val(14) is not None else None,
            "orders_other": int(get_val(15)) if get_val(15) is not None else None,
            # リピート情報
            "new_customers": int(get_val(29)) if get_val(29) is not None else None,
            "ec_site_buyers": int(get_val(31)) if get_val(31) is not None else None,
            "repeat_buyers": int(get_val(32)) if get_val(32) is not None else None,
            "repeat_single_month": int(get_val(33)) if get_val(33) is not None else None,
            "repeat_multi_month": int(get_val(34)) if get_val(34) is not None else None,
            # 返品・苦情
            "reshipping_count": int(get_val(39)) if get_val(39) is not None else None,
            "complaint_count": int(get_val(40)) if get_val(40) is not None else None,
            # 口コミ
            "positive_reviews": int(get_val(45)) if get_val(45) is not None else None,
            "negative_reviews": int(get_val(46)) if get_val(46) is not None else None,
            # コメント（各セクションのCol L）
            "comment_sales": get_comment(4),
            "comment_repeat": get_comment(29),
            "comment_complaint": get_comment(39),
            "comment_review": get_comment(45),
        }

        # 週次データ（Col E〜I: 第1〜5週）
        def build_weekly(mapping):
            """週次データJSONBを構築。全指標がNoneなら None を返す"""
            result = {}
            has_data = False
            for key, row in mapping.items():
                values = get_weekly(row)
                result[key] = values
                if any(v is not None for v in values):
                    has_data = True
            return result if has_data else None

        weekly_sales = build_weekly({
            "orders": 5,
            "sales": 8,
            "orders_kyushu": 11,
            "orders_chugoku_shikoku": 12,
            "orders_kansai": 13,
            "orders_kanto": 14,
            "orders_other": 15,
        })
        if weekly_sales is not None:
            data["weekly_sales"] = weekly_sales

        weekly_repeat = build_weekly({
            "new_customers": 29,
            "ec_site_buyers": 31,
            "repeat_buyers": 32,
            "repeat_single_month": 33,
            "repeat_multi_month": 34,
        })
        if weekly_repeat is not None:
            data["weekly_repeat"] = weekly_repeat

        weekly_complaint = build_weekly({
            "reshipping_count": 39,
            "complaint_count": 40,
        })
        if weekly_complaint is not None:
            data["weekly_complaint"] = weekly_complaint

        weekly_review = build_weekly({
            "positive_reviews": 45,
            "negative_reviews": 46,
        })
        if weekly_review is not None:
            data["weekly_review"] = weekly_review

        # Noneのキーを除外
        data = {k: v for k, v in data.items() if v is not None}

        if not data:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="アップロードするデータがありません"
            )

        await import_furusato_data(supabase, target_month, data)

        return FurusatoUploadResponse(
            success=True,
            message="ふるさと納税データのアップロードが完了しました",
            month=target_month.isoformat(),
            records_processed=1,
        )

    except HTTPException:
        raise
    except Exception as e:
        import traceback
        print(f"Furusato Excel upload error: {str(e)}")
        print(traceback.format_exc())
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
            detail=f"ファイルの処理中にエラーが発生しました: {str(e)}"
        )

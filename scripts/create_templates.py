"""
Excelテンプレート生成スクリプト

財務データ入力用・製造データ入力用のExcelテンプレートを生成する。

Usage:
    python create_templates.py --year 2025 --month 11
    python create_templates.py  # 現在の年月を使用
"""
import argparse
import calendar
from datetime import date, datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


# スタイル定義
HEADER_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
HEADER_FONT = Font(bold=True)
CENTER_ALIGNMENT = Alignment(horizontal="center", vertical="center")
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def create_financial_template(year: int, month: int, output_dir: Path) -> Path:
    """
    財務データ入力用Excelテンプレートを生成する

    Args:
        year: 対象年
        month: 対象月
        output_dir: 出力ディレクトリ

    Returns:
        生成されたファイルのパス
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "月次財務データ"

    # 列幅設定
    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 15
    ws.column_dimensions["C"].width = 8

    # ヘッダー情報
    ws["A1"] = "対象年月"
    ws["B1"] = f"{year}/{month:02d}/01"
    ws["A2"] = "データ区分"
    ws["B2"] = "実績"

    # 財務項目定義
    items = [
        ("■ 売上高", None, None, True),  # (項目名, セル参照, 単位, セクションヘッダーか)
        ("全社売上高", None, "円", False),
        ("店舗部門売上高", None, "円", False),
        ("通販部門売上高", None, "円", False),
        ("■ 原価・利益", None, None, True),
        ("売上原価", None, "円", False),
        ("売上総利益（粗利）", "=B6-B10", "円", False),  # 全社売上高 - 売上原価
        ("粗利率", "=IF(B6>0,B11/B6*100,0)", "%", False),  # 粗利 / 全社売上高 × 100
        ("■ 販管費", None, None, True),
        ("販管費合計", None, "円", False),
        ("うち人件費", None, "円", False),
        ("人件費率", "=IF(B6>0,B15/B6*100,0)", "%", False),  # 人件費 / 全社売上高 × 100
        ("その他経費", None, "円", False),
        ("■ 営業利益", None, None, True),
        ("営業利益", "=B11-B14", "円", False),  # 粗利 - 販管費合計
        ("営業利益率", "=IF(B6>0,B19/B6*100,0)", "%", False),  # 営業利益 / 全社売上高 × 100
        ("■ キャッシュフロー", None, None, True),
        ("営業キャッシュフロー", None, "円", False),
        ("投資キャッシュフロー", None, "円", False),
        ("財務キャッシュフロー", None, "円", False),
        ("フリーキャッシュフロー", "=B22+B23", "円", False),  # 営業CF + 投資CF
    ]

    row = 4
    for item_name, formula, unit, is_section in items:
        ws[f"A{row}"] = item_name
        if formula:
            ws[f"B{row}"] = formula
        if unit:
            ws[f"C{row}"] = unit

        # セクションヘッダーのスタイル
        if is_section:
            ws[f"A{row}"].font = HEADER_FONT
            ws[f"A{row}"].fill = HEADER_FILL
            ws[f"B{row}"].fill = HEADER_FILL
            ws[f"C{row}"].fill = HEADER_FILL

        row += 1

    # ファイル保存
    filename = f"financial_template_{year}{month:02d}.xlsx"
    filepath = output_dir / filename
    wb.save(filepath)
    print(f"財務テンプレート生成完了: {filepath}")
    return filepath


def create_manufacturing_template(year: int, month: int, output_dir: Path) -> Path:
    """
    製造データ入力用Excelテンプレートを生成する

    Args:
        year: 対象年
        month: 対象月
        output_dir: 出力ディレクトリ

    Returns:
        生成されたファイルのパス
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "日次製造データ"

    # 列幅設定
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 15
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 18
    ws.column_dimensions["F"].width = 15

    # ヘッダー情報
    ws["A1"] = "対象年月"
    ws["B1"] = f"{year}年{month:02d}月"

    # テーブルヘッダー
    headers = [
        "日付",
        "製造量(バット)",
        "製造量(個)",
        "出勤者数",
        "1人あたり製造量",
        "有給取得(時間)",
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER_ALIGNMENT
        cell.border = THIN_BORDER

    # 月の日数を取得
    _, days_in_month = calendar.monthrange(year, month)

    # データ行を生成
    data_start_row = 4
    for day in range(1, days_in_month + 1):
        row = data_start_row + day - 1
        day_date = date(year, month, day)

        # 日付
        ws.cell(row=row, column=1, value=day_date).border = THIN_BORDER

        # 製造量(バット) - 入力欄
        ws.cell(row=row, column=2).border = THIN_BORDER

        # 製造量(個) - 計算式: バット数 × 60
        cell_c = ws.cell(row=row, column=3, value=f"=B{row}*60")
        cell_c.border = THIN_BORDER

        # 出勤者数 - 入力欄
        ws.cell(row=row, column=4).border = THIN_BORDER

        # 1人あたり製造量 - 計算式
        cell_e = ws.cell(row=row, column=5, value=f"=IF(D{row}>0,B{row}/D{row},0)")
        cell_e.border = THIN_BORDER

        # 有給取得(時間) - 入力欄
        ws.cell(row=row, column=6).border = THIN_BORDER

    # 合計行
    sum_row = data_start_row + days_in_month
    ws.cell(row=sum_row, column=1, value="合計").font = HEADER_FONT
    ws.cell(row=sum_row, column=1).fill = HEADER_FILL
    ws.cell(row=sum_row, column=1).border = THIN_BORDER

    # 各列の合計
    for col in range(2, 7):
        col_letter = get_column_letter(col)
        if col == 5:  # 1人あたり製造量は平均計算
            formula = f"=IF(D{sum_row}>0,B{sum_row}/D{sum_row},0)"
        else:
            formula = f"=SUM({col_letter}{data_start_row}:{col_letter}{sum_row-1})"

        cell = ws.cell(row=sum_row, column=col, value=formula)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER

    # ファイル保存
    filename = f"manufacturing_template_{year}{month:02d}.xlsx"
    filepath = output_dir / filename
    wb.save(filepath)
    print(f"製造テンプレート生成完了: {filepath}")
    return filepath


def main():
    """メイン関数"""
    parser = argparse.ArgumentParser(description="財務・製造データ用Excelテンプレートを生成")
    parser.add_argument("--year", type=int, default=None, help="対象年（デフォルト: 現在の年）")
    parser.add_argument("--month", type=int, default=None, help="対象月（デフォルト: 現在の月）")
    parser.add_argument(
        "--output",
        type=str,
        default=None,
        help="出力ディレクトリ（デフォルト: backend/templates/）",
    )

    args = parser.parse_args()

    # デフォルト値設定
    now = datetime.now()
    year = args.year or now.year
    month = args.month or now.month

    # 出力ディレクトリ
    if args.output:
        output_dir = Path(args.output)
    else:
        # スクリプトの親ディレクトリ/templates/
        script_dir = Path(__file__).parent
        output_dir = script_dir.parent / "templates"

    output_dir.mkdir(parents=True, exist_ok=True)

    print(f"対象年月: {year}年{month}月")
    print(f"出力先: {output_dir}")
    print("-" * 40)

    # テンプレート生成
    create_financial_template(year, month, output_dir)
    create_manufacturing_template(year, month, output_dir)

    print("-" * 40)
    print("完了")


if __name__ == "__main__":
    main()
